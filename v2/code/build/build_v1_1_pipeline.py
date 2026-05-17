"""
v1.1 — D-4 (channel_keywords v0.2) + D-5 (subtype 활성채널 재매핑)
================================================================

D-4: C4 키워드 정제
  제거: 법안, 승인, 심사, 발표, 추진, 대책 (정책 일반 행정어)
  유지: 규제강화, 불확실성, 리스크, 제재, 제한, 금지, 위축, 압박, 갈등

D-5: subtype 활성 채널 재매핑
  A_tax_support              → [C1]              (C5 제외 — D-2에서 K칩스법 C5 0% 활성)
  B_restriction_negative     → [C2, C4]
  B_market_stabilizer        → [C5, C6]
  C_industrial_support       → [C3, C6]
  D_crisis_relief            → [C5, C4, C6]
  D_crisis_shock             → [C4, C6]
  E_foreign_benefit          → [C2, C3]
  E_foreign_shock            → [C3, C4]

출력:
  channel_keywords_v0.2.json
  policy_cards_v1.1.csv
  policy_hypotheses_v1.1.csv
"""
import json, csv
import openpyxl

# ═══════════════════════════════════════
# D-4: channel_keywords v0.2
# ═══════════════════════════════════════
with open('channel_keywords.json', encoding='utf-8') as f:
    ck = json.load(f)

# C4 정제: neutral_anchors에서 일반 행정어 제거
C4_REMOVE_FROM_NEUTRAL = {"법", "법안", "제도", "정책", "행정", "심사", "승인", "인허가", "발표", "추진", "대책"}
# polarity_keywords.supports_higher_regulation 보강 (위축, 갈등 추가 — 정책 유발 부정 효과)
C4_ADD_TO_HIGHER = ["위축", "갈등"]
# polarity_keywords.supports_lower_regulation 보강 (안정, 신뢰회복 추가)
C4_ADD_TO_LOWER = ["안정", "신뢰회복"]

ck['version'] = '0.2'
ck['updated'] = '2026-05-16'
ck['changelog_v0.2'] = [
    "D-4 보정: C4 neutral_anchors에서 정책 일반 행정어 제거",
    f"  removed from C4.neutral_anchors: {sorted(C4_REMOVE_FROM_NEUTRAL)}",
    f"  added to C4.polarity_keywords.supports_higher_regulation: {C4_ADD_TO_HIGHER}",
    f"  added to C4.polarity_keywords.supports_lower_regulation: {C4_ADD_TO_LOWER}",
    "근거: D-2 sanity에서 C4 활성률 30~100% 과활성 — 정책 일반 행정어가 채널 표지로 부적합",
]

c4 = ck['channels']['C4']
c4['neutral_anchors'] = [w for w in c4['neutral_anchors'] if w not in C4_REMOVE_FROM_NEUTRAL]
c4['polarity_keywords']['supports_higher_regulation'].extend(C4_ADD_TO_HIGHER)
c4['polarity_keywords']['supports_lower_regulation'].extend(C4_ADD_TO_LOWER)

with open('channel_keywords_v0.2.json', 'w', encoding='utf-8') as f:
    json.dump(ck, f, ensure_ascii=False, indent=2)

print(f"[1/4] channel_keywords_v0.2.json 저장")
print(f"      C4 neutral_anchors: {len(c4['neutral_anchors'])}개 (정책 일반어 제거됨)")
print(f"      C4 higher_regulation: {len(c4['polarity_keywords']['supports_higher_regulation'])}개")

# ═══════════════════════════════════════
# D-5: subtype 활성 채널 재매핑
# ═══════════════════════════════════════
SUBTYPE_ACTIVE_V1_1 = {
    'A_tax_support':          [('C1', '+')],                          # C5 제외
    'B_restriction_negative': [('C2', '-'), ('C4', '-')],
    'B_market_stabilizer':    [('C5', '+'), ('C6', '+')],
    'C_industrial_support':   [('C3', '+'), ('C6', '+')],
    'D_crisis_relief':        [('C5', '+'), ('C4', '+'), ('C6', '+')],  # C4는 완화방향
    'D_crisis_shock':         [('C4', '-'), ('C6', '-')],
    'E_foreign_benefit':      [('C2', '+'), ('C3', '+')],
    'E_foreign_shock':        [('C3', '-'), ('C4', '-')],              # C2 제외 (수요는 일관성 약함)
}

# v1.0 카드 로드 + 매핑 갱신
with open('policy_cards_v1.0.csv', encoding='utf-8-sig') as f:
    cards = list(csv.DictReader(f))

for c in cards:
    sub = c['type_subtype']
    # 기존 C1~C6 초기화
    for cc in ['C1','C2','C3','C4','C5','C6']:
        c[cc] = ''
    # 새 매핑 적용
    chan_signs = {}
    for ch, sign in SUBTYPE_ACTIVE_V1_1.get(sub, []):
        c[ch] = 'ON'
        chan_signs[ch] = sign
    c['channel_signs'] = str(chan_signs)
    c['status'] = 'v1.1_frozen'
    c['frozen_at'] = '2026-05-16'

keys_v1 = ['pn','industry','alias','date','master_type','type_subtype',
           'D_i_initial','D_i_source','D_i_final_car','k_i',
           'C1','C2','C3','C4','C5','C6','channel_signs',
           'rationale','review_note','status','frozen_at','review_status']

with open('policy_cards_v1.1.csv', 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=keys_v1, extrasaction='ignore')
    w.writeheader()
    for c in cards:
        w.writerow({k: c.get(k, '') for k in keys_v1})

print(f"\n[2/4] policy_cards_v1.1.csv 저장 ({len(cards)}건)")

# subtype별 분포 확인
from collections import Counter
print(f"\n      subtype별 정책 수:")
sub_counter = Counter(c['type_subtype'] for c in cards)
for s, n in sorted(sub_counter.items(), key=lambda x: -x[1]):
    chans = ','.join(ch for ch, _ in SUBTYPE_ACTIVE_V1_1.get(s, []))
    print(f"        {s:<28s} {n:2d}건  active=[{chans}]")

# ═══════════════════════════════════════
# 가설 v1.1 재생성 (build_hypothesis 동일, 카드만 갱신)
# ═══════════════════════════════════════
SUBTYPE_TEMPLATES = {
    'A_tax_support': {
        'pos': "'{policy}' 정책이 세부담 감소와 자금 여력 개선을 통해 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 정책이 세부담 증가나 정책 효과 부진으로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'B_restriction_negative': {
        'pos': "'{policy}' 정책이 시장 정상화와 신뢰 회복을 통해 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 정책이 매출 위축과 정책 불확실성 확대로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'B_market_stabilizer': {
        'pos': "'{policy}' 정책이 시장 안정·유동성 회복·투자심리 개선을 통해 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 정책이 거래 위축과 시장 왜곡으로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'C_industrial_support': {
        'pos': "'{policy}' 정책이 설비투자 확대·인프라 구축·시장 신뢰 강화를 통해 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 정책이 실행 지연과 정책 효과 부진으로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'D_crisis_relief': {
        # v1.1: C4 명시 — 불확실성 "완화" 방향 (shock과 구분)
        'pos': "'{policy}' 정책이 유동성 공급·시장 불안 완화·신뢰 회복을 통해 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 정책이 정책 효과 미흡과 추가 위기 우려 확산으로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'D_crisis_shock': {
        # v1.1: C4 명시 — 불확실성 "확대" 방향
        'pos': "'{policy}' 충격이 시장 정화와 점진적 신뢰 회복으로 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 충격이 패닉 확산·신뢰 훼손·불확실성과 리스크 확대로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'E_foreign_benefit': {
        'pos': "'{policy}'이(가) 한국 기업의 수출 수혜와 공급망 우위를 통해 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}'이(가) 글로벌 경쟁 격화와 기대 미달로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
    'E_foreign_shock': {
        'pos': "'{policy}' 충격에 대해 한국 기업의 대안 시장 확보와 국산화 가속으로 {stocks} 등 {industry}의 주가에 상승 압력을 만든다",
        'neg': "'{policy}' 충격이 공급망 위축·정책 불확실성 확대로 {stocks} 등 {industry}의 주가에 하락 압력을 만든다",
    },
}
INDUSTRY_MAP = {
    '반도체': '반도체 업종', '금융/증권': '금융 업종', '부동산/건설': '건설 업종',
    '바이오/제약': '제약·바이오 업종', '2차전지/에너지': '2차전지 업종',
}

def build_hypothesis(card, top_stocks):
    sub = card['type_subtype']
    if sub not in SUBTYPE_TEMPLATES: return None, None, None
    tpl = SUBTYPE_TEMPLATES[sub]
    industry = INDUSTRY_MAP.get(card['industry'], card['industry'])
    stocks = ', '.join([s.strip() for s in top_stocks.split(',')[:2]
                        if s.strip() and s.strip() != '전 종목']) or industry
    pos = tpl['pos'].format(policy=card['alias'], stocks=stocks, industry=industry)
    neg = tpl['neg'].format(policy=card['alias'], stocks=stocks, industry=industry)
    neutral = f"이 기사는 '{card['alias']}' 정책과 무관하거나 정책의 주가 영향이 아닌 단순 사실 전달이다"
    if card['D_i_initial'] == '+': return pos, neg, neutral
    return neg, pos, neutral

# top_stocks
wb = openpyxl.load_workbook('50_정책마스터.xlsx', read_only=True, data_only=True)
ws = wb['1_50정책_마스터']
TOP_STOCKS = {}
for r in list(ws.iter_rows(values_only=True))[3:]:
    if not r or not r[0]: continue
    TOP_STOCKS[int(r[0])] = r[9] or ''
wb.close()

hyps = []
for c in cards:
    pn = int(c['pn'])
    sup, con, neu = build_hypothesis(c, TOP_STOCKS.get(pn, ''))
    hyps.append({
        'pn': pn, 'alias': c['alias'], 'industry': c['industry'],
        'type_subtype': c['type_subtype'], 'D_i_initial': c['D_i_initial'],
        'active_channels': ','.join([cc for cc in ['C1','C2','C3','C4','C5','C6'] if c.get(cc)=='ON']),
        'hypothesis_support': sup, 'hypothesis_contradict': con, 'hypothesis_neutral': neu,
        'status': 'v1.1_frozen',
    })

with open('policy_hypotheses_v1.1.csv', 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(hyps[0].keys()))
    w.writeheader()
    w.writerows(hyps)

print(f"\n[3/4] policy_hypotheses_v1.1.csv 저장 ({len(hyps)}건)")
print(f"      D_crisis_relief C4 가설 — '시장 불안 완화'로 명시 (shock과 구분)")
print(f"      D_crisis_shock C4 가설 — '불확실성과 리스크 확대'로 명시")

# 14_PolicyCard 시트 갱신 (v1.1로)
wb = openpyxl.load_workbook('50_정책마스터.xlsx')
if '14_PolicyCard' in wb.sheetnames:
    del wb['14_PolicyCard']
ws_new = wb.create_sheet('14_PolicyCard')
ws_new.append(keys_v1)
for c in cards:
    ws_new.append([c.get(k, '') for k in keys_v1])
wb.save('50_정책마스터.xlsx')
print(f"\n[4/4] 14_PolicyCard 시트 v1.1로 갱신")
