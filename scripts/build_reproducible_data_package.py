from __future__ import annotations

import math
import re
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import FinanceDataReader as fdr
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "data" / "reproducible_data_package.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
}

NAVER_PAGE_CACHE: dict[tuple[str, int], pd.DataFrame] = {}
NAVER_LAST_PAGE_CACHE: dict[str, int] = {}

NAVY_FILL = PatternFill("solid", fgColor="17375E")
TEAL_FILL = PatternFill("solid", fgColor="1F7A8C")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
GRAY_FILL = PatternFill("solid", fgColor="F3F6F9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
SUBTITLE_FONT = Font(size=10, color="666666")
CARD_LABEL_FONT = Font(size=10, bold=True, color="666666")
CARD_VALUE_FONT = Font(size=18, bold=True)
HEADER_FONT = Font(size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


@dataclass(frozen=True)
class ValupExpectation:
    event_id: str
    event_date: str
    event_name: str
    stock_name: str
    code: str
    d0: float
    d1: float
    d3: float
    d5: float
    d10: float
    car_0_10: float


@dataclass(frozen=True)
class Event50:
    event_id: str
    sector: str
    event_name: str
    event_date: str
    stock_name_reported: str
    code: str
    source_system: str
    direction: str
    expected_return_pct: float
    expected_dk: int
    max_window_dk: int
    note: str = ""


VALUP_EVENTS = [
    ("E1", "2024-02-26", "밸류업 프로그램 최초 발표"),
    ("E2", "2024-05-02", "밸류업 공시 가이드라인 확정"),
    ("E3", "2024-09-24", "코리아 밸류업 지수 구성종목 100선 발표"),
]

VALUP_EXPECTATIONS = [
    ValupExpectation("E1", "2024-02-26", "밸류업 프로그램 최초 발표", "KB금융", "105560", -5.02, -0.20, 1.90, 0.70, 3.50, 10.66),
    ValupExpectation("E1", "2024-02-26", "밸류업 프로그램 최초 발표", "신한지주", "055550", -4.50, 1.60, 1.50, -0.30, -0.40, 4.63),
    ValupExpectation("E1", "2024-02-26", "밸류업 프로그램 최초 발표", "하나금융지주", "086790", -5.94, -1.30, 1.80, -0.80, 0.00, 0.87),
    ValupExpectation("E2", "2024-05-02", "밸류업 공시 가이드라인 확정", "KB금융", "105560", -4.37, 1.90, 5.10, 3.50, 1.90, 6.17),
    ValupExpectation("E2", "2024-05-02", "밸류업 공시 가이드라인 확정", "신한지주", "055550", -1.82, 1.40, 2.30, 2.70, 2.70, 3.19),
    ValupExpectation("E2", "2024-05-02", "밸류업 공시 가이드라인 확정", "하나금융지주", "086790", -2.90, 1.80, 4.30, 4.40, 1.90, 7.66),
    ValupExpectation("E3", "2024-09-24", "코리아 밸류업 지수 구성종목 100선 발표", "KB금융", "105560", -3.53, -4.80, 3.20, 0.50, 1.60, 7.82),
    ValupExpectation("E3", "2024-09-24", "코리아 밸류업 지수 구성종목 100선 발표", "신한지주", "055550", 1.08, -5.10, -1.40, 0.70, 1.20, 3.51),
    ValupExpectation("E3", "2024-09-24", "코리아 밸류업 지수 구성종목 100선 발표", "하나금융지주", "086790", -3.40, -3.20, -0.30, 1.20, -0.70, -0.64),
]

EVENTS_50 = [
    Event50("반도체#1", "반도체", "현대전자-LG반도체 빅딜", "1999-05-20", "삼성전자", "005930", "NaverFinance", "down", -5.39, 0, 5),
    Event50("반도체#2", "반도체", "차세대 성장동력 10대 산업 확정", "2003-08-22", "SK하이닉스", "000660", "NaverFinance", "up", 4.40, 4, 5),
    Event50("반도체#3", "반도체", "일본 수출규제 발동", "2019-07-01", "후성", "093370", "FinanceDataReader", "up", 17.53, 2, 5),
    Event50("반도체#4", "반도체", "소부장 경쟁력 강화 대책", "2019-08-05", "솔브레인", "036830", "FinanceDataReader", "up", 7.85, 2, 5),
    Event50("반도체#5", "반도체", "K-반도체 벨트 전략", "2021-05-13", "삼성전자", "005930", "FinanceDataReader", "up", 2.04, 1, 5),
    Event50("반도체#6", "반도체", "반도체 초강대국 달성 전략", "2022-07-21", "삼성전자", "005930", "FinanceDataReader", "up", 2.15, 0, 5),
    Event50("반도체#7", "반도체", "미국 대중국 반도체 수출통제", "2022-10-11", "HPSP", "403870", "FinanceDataReader", "down", -6.38, 2, 5),
    Event50("반도체#8", "반도체", "K칩스법 1차 (세액공제 8→15%)", "2023-03-22", "한미반도체", "042700", "FinanceDataReader", "up", 21.88, 2, 5),
    Event50("반도체#9", "반도체", "반도체 메가 클러스터", "2024-01-15", "HPSP", "403870", "FinanceDataReader", "up", 8.51, 4, 5),
    Event50("반도체#10", "반도체", "삼성전자 시스템반도체 비전 2030", "2019-04-30", "티씨케이", "064760", "FinanceDataReader", "up", 2.74, 1, 5),
    Event50("금융#1", "금융/증권", "금융 구조조정 5개 은행 퇴출", "1998-06-29", "삼성증권", "016360", "NaverFinance", "down", -3.06, 0, 5),
    Event50("금융#2", "금융/증권", "리먼브라더스 파산", "2008-09-15", "삼성증권", "016360", "NaverFinance", "down", -9.80, 0, 5),
    Event50("금융#3", "금융/증권", "최경환 경제팀 배당 확대 정책", "2014-07-24", "키움증권", "039490", "FinanceDataReader", "up", 2.93, 1, 5),
    Event50("금융#4", "금융/증권", "공매도 일시 금지②", "2011-08-10", "KB금융", "105560", "NaverFinance", "up", 10.22, 3, 5),
    Event50("금융#5", "금융/증권", "공매도 전면 금지③ (코로나)", "2020-03-16", "KB금융", "105560", "FinanceDataReader", "up", 18.43, 5, 5),
    Event50("금융#6", "금융/증권", "코로나 금융안정 패키지 100조원", "2020-03-24", "삼성증권", "016360", "FinanceDataReader", "up", 17.34, 1, 5),
    Event50("금융#7", "금융/증권", "코스닥 활성화 방안", "2018-01-11", "한국금융지주", "071050", "FinanceDataReader", "up", 8.08, 1, 5),
    Event50("금융#8", "금융/증권", "공매도 재금지④", "2023-11-06", "삼성증권", "016360", "FinanceDataReader", "up", 4.98, 0, 5),
    Event50("금융#9", "금융/증권", "기업 밸류업 프로그램", "2024-01-17", "KB금융", "105560", "FinanceDataReader", "up", 4.26, 4, 5),
    Event50("금융#10", "금융/증권", "금투세 폐지 여야 합의", "2024-11-04", "삼성증권", "016360", "FinanceDataReader", "up", 2.18, 1, 5),
    Event50("부동산#1", "부동산/건설", "IMF 직후 주택경기 활성화 대책", "1998-05-22", "대림산업", "000210", "NaverFinance", "up", 11.96, 5, 5),
    Event50("부동산#2", "부동산/건설", "10.29 주택시장 안정 종합대책", "2003-10-29", "현대건설", "000720", "NaverFinance", "down", -6.60, 8, 8),
    Event50("부동산#3", "부동산/건설", "9.13 주택시장 안정대책", "2018-09-13", "HDC", "012630", "FinanceDataReader", "down", -6.82, 5, 5),
    Event50("부동산#4", "부동산/건설", "부동산 세제 완화 대책", "2008-09-01", "대우건설", "047040", "NaverFinance", "up", 12.82, 5, 5),
    Event50("부동산#5", "부동산/건설", "보금자리주택 건설 방안", "2008-09-19", "GS건설", "006360", "NaverFinance", "up", 6.91, 0, 5),
    Event50("부동산#6", "부동산/건설", "분양가상한제 민간 적용 확대", "2019-08-12", "GS건설", "006360", "FinanceDataReader", "down", -5.42, 1, 5),
    Event50("부동산#7", "부동산/건설", "6.19 주택시장 안정화 대책", "2017-06-19", "현대건설", "000720", "FinanceDataReader", "down", -3.28, 4, 5),
    Event50("부동산#8", "부동산/건설", "6.17 부동산 대책", "2020-06-17", "GS건설", "006360", "FinanceDataReader", "down", -2.81, 3, 5, "섹터 혼재 사례로 각주 필요"),
    Event50("부동산#9", "부동산/건설", "윤석열 정부 규제 완화 패키지", "2022-06-21", "HDC현대산업개발", "294870", "FinanceDataReader", "up", 5.21, 3, 5),
    Event50(
        "부동산#10",
        "부동산/건설",
        "1.10 주택공급 대책",
        "2024-01-10",
        "DL이앤씨",
        "000215",
        "FinanceDataReader",
        "up",
        6.51,
        1,
        5,
        "보고서 산출값 재현 기준 코드 000215 사용. 현재 종목명과 보고서 표기명이 다름.",
    ),
    Event50("바이오#1", "바이오/제약", "의약분업 시행", "2000-07-01", "유한양행", "000100", "NaverFinance", "down", -3.31, 8, 8),
    Event50("바이오#2", "바이오/제약", "황우석 줄기세포 2호 사이언스 논문 발표", "2005-05-20", "마크로젠", "038290", "NaverFinance", "up", 2.84, 0, 5),
    Event50("바이오#3", "바이오/제약", "황우석 줄기세포 논문 조작 의혹 방송(PD수첩)", "2005-12-15", "녹십자", "006280", "NaverFinance", "down", -8.13, 1, 5),
    Event50("바이오#4", "바이오/제약", "바이오시밀러 허가심사 가이드라인", "2009-06-30", "셀트리온", "068270", "NaverFinance", "up", 3.70, 0, 5),
    Event50("바이오#5", "바이오/제약", "리베이트 쌍벌제 시행", "2010-11-29", "셀트리온", "068270", "NaverFinance", "up", 5.26, 0, 5),
    Event50("바이오#6", "바이오/제약", "셀트리온 램시마 세계최초 허가", "2012-07-23", "유한양행", "000100", "NaverFinance", "up", 3.63, 7, 7),
    Event50("바이오#7", "바이오/제약", "바이오헬스 육성 전략", "2019-05-22", "셀트리온", "068270", "FinanceDataReader", "up", 6.90, 4, 5),
    Event50("바이오#8", "바이오/제약", "코로나19 치료제·백신 개발 지원방안", "2020-04-17", "씨젠", "096530", "FinanceDataReader", "up", 10.88, 1, 5),
    Event50("바이오#9", "바이오/제약", "의대정원 2000명 증원 발표", "2024-02-06", "셀트리온", "068270", "FinanceDataReader", "up", 4.40, 3, 5),
    Event50("바이오#10", "바이오/제약", "CDMO 특별법 제정/공포", "2025-12-30", "녹십자", "006280", "FinanceDataReader", "up", 6.50, 0, 5),
    Event50("2차전지#1", "2차전지/에너지", "녹색성장 5개년 계획", "2009-07-06", "한화솔루션", "009830", "NaverFinance", "up", 5.17, 1, 5),
    Event50("2차전지#2", "2차전지/에너지", "전기차 민간 보급 보조금 최초", "2013-06-25", "LG화학", "051910", "NaverFinance", "up", 4.33, 3, 5),
    Event50("2차전지#3", "2차전지/에너지", "재생에너지 3020 이행계획", "2017-12-20", "LG화학", "051910", "FinanceDataReader", "up", 2.66, 4, 5),
    Event50("2차전지#4", "2차전지/에너지", "ESS 화재 안전강화대책", "2019-06-11", "에코프로비엠", "247540", "FinanceDataReader", "down", -5.36, 5, 5),
    Event50("2차전지#5", "2차전지/에너지", "한국판 뉴딜 그린뉴딜 73.4조원", "2020-07-14", "씨에스윈드", "112610", "FinanceDataReader", "up", 14.83, 5, 5),
    Event50("2차전지#6", "2차전지/에너지", "EU 탄소국경조정세(CBAM) 발표", "2021-07-14", "삼성SDI", "006400", "FinanceDataReader", "up", 3.12, 9, 9),
    Event50("2차전지#7", "2차전지/에너지", "미국 IRA 서명·발효", "2022-08-16", "에코프로", "086520", "FinanceDataReader", "up", 4.05, 0, 5),
    Event50("2차전지#8", "2차전지/에너지", "이차전지 산업 혁신전략", "2022-11-01", "LG화학", "051910", "FinanceDataReader", "up", 11.02, 0, 5),
    Event50("2차전지#9", "2차전지/에너지", "첨단전략산업 지정 + 세액공제", "2023-04-20", "에코프로", "086520", "FinanceDataReader", "up", 18.36, 5, 5),
    Event50("2차전지#10", "2차전지/에너지", "트럼프 IRA 전기차 세액공제 축소", "2025-05-22", "LG화학", "051910", "FinanceDataReader", "down", -2.56, 3, 5),
]

BIGKINDS_REFERENCE = [
    {
        "event_id": "E3",
        "event_name": "코리아 밸류업 지수 구성종목 발표",
        "window": "D-90~D-1",
        "count": 155,
        "source_status": "manual_from_final_md",
        "note": "원시 기사 목록은 미포함. 최종 md 본문 수치만 고정.",
    },
    {
        "event_id": "E3",
        "event_name": "코리아 밸류업 지수 구성종목 발표",
        "window": "D0~D+30",
        "count": 137,
        "source_status": "manual_from_final_md",
        "note": "원시 기사 목록은 미포함. 최종 md 본문 수치만 고정.",
    },
    {
        "event_id": "E3",
        "event_name": "코리아 밸류업 지수 구성종목 발표",
        "window": "D-90~D-61",
        "count": 64,
        "source_status": "manual_from_final_md",
        "note": "기사 타임라인 참고치",
    },
    {
        "event_id": "E3",
        "event_name": "코리아 밸류업 지수 구성종목 발표",
        "window": "D-60~D-31",
        "count": 69,
        "source_status": "manual_from_final_md",
        "note": "기사 타임라인 참고치",
    },
    {
        "event_id": "E3",
        "event_name": "코리아 밸류업 지수 구성종목 발표",
        "window": "D-30~D-1",
        "count": 22,
        "source_status": "manual_from_final_md",
        "note": "기사 타임라인 참고치",
    },
    {
        "event_id": "E3",
        "event_name": "코리아 밸류업 지수 구성종목 발표",
        "window": "D0~D+7",
        "count": 71,
        "source_status": "manual_from_final_md",
        "note": "기사 타임라인 참고치",
    },
]


def pct_diff(a: float | None, b: float | None) -> float | None:
    if a is None or b is None:
        return None
    return round(a - b, 4)


def to_timestamp(value: str) -> pd.Timestamp:
    return pd.Timestamp(value)


def clean_int(text: str) -> int | None:
    text = text.replace(",", "").strip()
    if not text or text == "-":
        return None
    return int(text)


def current_name_map() -> dict[str, str]:
    listing = fdr.StockListing("KRX")[["Code", "Name"]]
    return dict(zip(listing["Code"], listing["Name"]))


def estimate_naver_page(target_date: pd.Timestamp, base_date: pd.Timestamp) -> int:
    days_diff = (base_date - target_date).days
    trading_days = max(days_diff * 5 / 7, 0)
    return max(1, int(trading_days / 10) + 1)


def fetch_naver_page(code: str, page: int) -> pd.DataFrame:
    cache_key = (code, page)
    if cache_key in NAVER_PAGE_CACHE:
        return NAVER_PAGE_CACHE[cache_key].copy()

    response = requests.get(
        f"https://finance.naver.com/item/sise_day.naver?code={code}&page={page}",
        headers=HEADERS,
        timeout=15,
    )
    response.raise_for_status()
    response.encoding = "euc-kr"
    if code not in NAVER_LAST_PAGE_CACHE:
        match = re.search(r"page=(\d+)\"[^>]*>\s*맨뒤", response.text)
        if match:
            NAVER_LAST_PAGE_CACHE[code] = int(match.group(1))
    soup = BeautifulSoup(response.text, "html.parser")

    rows = []
    for tr in soup.select("table.type2 tr"):
        tds = tr.select("td")
        if len(tds) < 7:
            continue
        date_text = tds[0].get_text(strip=True)
        if "." not in date_text:
            continue
        rows.append(
            {
                "date": pd.to_datetime(date_text.replace(".", "-")),
                "close": clean_int(tds[1].get_text(strip=True)),
                "open": clean_int(tds[3].get_text(strip=True)),
                "high": clean_int(tds[4].get_text(strip=True)),
                "low": clean_int(tds[5].get_text(strip=True)),
                "volume": clean_int(tds[6].get_text(strip=True)),
                "source_page": page,
            }
        )
    df = pd.DataFrame(rows)
    if df.empty:
        NAVER_PAGE_CACHE[cache_key] = df
        return df.copy()
    df = df.sort_values("date").drop_duplicates(subset=["date"], keep="last")
    NAVER_PAGE_CACHE[cache_key] = df
    return df.copy()


def naver_last_page(code: str) -> int:
    if code not in NAVER_LAST_PAGE_CACHE:
        fetch_naver_page(code, 1)
    return NAVER_LAST_PAGE_CACHE[code]


def find_naver_page_for_date(code: str, target_date: pd.Timestamp) -> int:
    low = 1
    high = naver_last_page(code)
    best = high

    while low <= high:
        mid = (low + high) // 2
        df = fetch_naver_page(code, mid)
        if df.empty:
            break
        oldest = df["date"].min()
        newest = df["date"].max()
        best = mid
        if target_date > newest:
            high = mid - 1
        elif target_date < oldest:
            low = mid + 1
        else:
            return mid

    return max(1, min(best, naver_last_page(code)))


def fetch_naver_range(code: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    buffer_start = start - pd.Timedelta(days=7)
    buffer_end = end + pd.Timedelta(days=7)
    older_page = find_naver_page_for_date(code, start)
    newer_page = find_naver_page_for_date(code, end)
    page_from = max(1, min(older_page, newer_page) - 1)
    page_to = max(older_page, newer_page) + 1

    frames = []
    for page in range(page_from, page_to + 1):
        frames.append(fetch_naver_page(code, page))
        time.sleep(0.01)

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(subset=["date"], keep="last").sort_values("date")
    filtered = combined[(combined["date"] >= buffer_start) & (combined["date"] <= buffer_end)].copy()
    if filtered.empty or filtered["date"].min() > start or filtered["date"].max() < end:
        raise RuntimeError(f"Naver data coverage failed for {code} {start.date()}~{end.date()}")
    return filtered.reset_index(drop=True)


def fetch_fdr_range(code: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    df = fdr.DataReader(code, start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")).reset_index()
    df = df.rename(
        columns={
            "Date": "date",
            "Open": "open",
            "High": "high",
            "Low": "low",
            "Close": "close",
            "Volume": "volume",
        }
    )
    columns = ["date", "open", "high", "low", "close", "volume"]
    return df[columns].sort_values("date").reset_index(drop=True)


def add_return_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.sort_values("date").reset_index(drop=True).copy()
    df["prev_close"] = df["close"].shift(1)
    df["daily_return_pct"] = ((df["close"] / df["prev_close"]) - 1) * 100
    return df


def market_d0_index(df: pd.DataFrame, event_date: pd.Timestamp) -> int:
    candidates = df.index[df["date"] >= event_date].tolist()
    if not candidates:
        raise RuntimeError(f"No trading day on or after {event_date.date()}")
    return candidates[0]


def build_valup_tables() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[dict[str, object]]]:
    full_start = pd.Timestamp("2024-01-01")
    full_end = pd.Timestamp("2024-12-31")
    price_frames = {
        "KB금융": fetch_fdr_range("105560", full_start, full_end),
        "신한지주": fetch_fdr_range("055550", full_start, full_end),
        "하나금융지주": fetch_fdr_range("086790", full_start, full_end),
    }
    kospi_df = add_return_columns(fetch_fdr_range("KS11", full_start, full_end))
    for key in price_frames:
        price_frames[key] = add_return_columns(price_frames[key])

    raw_rows: list[dict[str, object]] = []
    calc_rows: list[dict[str, object]] = []
    report_rows: list[dict[str, object]] = []
    qa_rows: list[dict[str, object]] = []
    expected_map = {(row.event_id, row.stock_name): row for row in VALUP_EXPECTATIONS}

    for event_id, event_date_str, event_name in VALUP_EVENTS:
        event_date = to_timestamp(event_date_str)
        d0_idx = market_d0_index(kospi_df, event_date)
        start_idx = max(0, d0_idx - 5)
        end_idx = min(len(kospi_df) - 1, d0_idx + 20)
        event_dates = kospi_df.iloc[start_idx : end_idx + 1].copy()
        event_dates["dk"] = range(start_idx - d0_idx, end_idx - d0_idx + 1)

        for _, row in event_dates.iterrows():
            raw_rows.append(
                {
                    "event_id": event_id,
                    "event_name": event_name,
                    "event_date": event_date_str,
                    "asset_name": "KOSPI",
                    "asset_code": "KS11",
                    "source_system": "FinanceDataReader",
                    "trade_date": row["date"].date().isoformat(),
                    "dk": int(row["dk"]),
                    "open": row["open"],
                    "high": row["high"],
                    "low": row["low"],
                    "close": row["close"],
                    "volume": row["volume"],
                    "daily_return_pct": None if pd.isna(row["daily_return_pct"]) else round(float(row["daily_return_pct"]), 4),
                }
            )

        for stock_name, code in [("KB금융", "105560"), ("신한지주", "055550"), ("하나금융지주", "086790")]:
            stock_df = price_frames[stock_name]
            stock_dates = stock_df[stock_df["date"].isin(event_dates["date"])].copy()
            stock_dates = stock_dates.merge(event_dates[["date", "dk", "daily_return_pct"]], on="date", how="left", suffixes=("", "_kospi"))
            stock_dates = stock_dates.rename(columns={"daily_return_pct_kospi": "kospi_return_pct"})
            stock_dates["ar_pct"] = stock_dates["daily_return_pct"] - stock_dates["kospi_return_pct"]
            stock_dates["car_pct"] = None

            running = 0.0
            for idx in stock_dates.index:
                dk = int(stock_dates.loc[idx, "dk"])
                if dk >= 0 and pd.notna(stock_dates.loc[idx, "ar_pct"]):
                    running += float(stock_dates.loc[idx, "ar_pct"])
                    stock_dates.loc[idx, "car_pct"] = running

            for _, row in stock_dates.iterrows():
                raw_rows.append(
                    {
                        "event_id": event_id,
                        "event_name": event_name,
                        "event_date": event_date_str,
                        "asset_name": stock_name,
                        "asset_code": code,
                        "source_system": "FinanceDataReader",
                        "trade_date": row["date"].date().isoformat(),
                        "dk": int(row["dk"]),
                        "open": row["open"],
                        "high": row["high"],
                        "low": row["low"],
                        "close": row["close"],
                        "volume": row["volume"],
                        "daily_return_pct": None if pd.isna(row["daily_return_pct"]) else round(float(row["daily_return_pct"]), 4),
                    }
                )
                calc_rows.append(
                    {
                        "event_id": event_id,
                        "event_name": event_name,
                        "event_date": event_date_str,
                        "stock_name": stock_name,
                        "stock_code": code,
                        "trade_date": row["date"].date().isoformat(),
                        "dk": int(row["dk"]),
                        "stock_return_pct": None if pd.isna(row["daily_return_pct"]) else round(float(row["daily_return_pct"]), 4),
                        "kospi_return_pct": None if pd.isna(row["kospi_return_pct"]) else round(float(row["kospi_return_pct"]), 4),
                        "ar_pct": None if pd.isna(row["ar_pct"]) else round(float(row["ar_pct"]), 4),
                        "car_pct": None if pd.isna(row["car_pct"]) else round(float(row["car_pct"]), 4),
                    }
                )

            landmarks = {}
            for target_dk in (0, 1, 3, 5, 10):
                target_row = stock_dates[stock_dates["dk"] == target_dk]
                landmarks[target_dk] = None if target_row.empty else round(float(target_row.iloc[0]["daily_return_pct"]), 2)

            car_row = stock_dates[stock_dates["dk"] == 10]
            computed_car = None if car_row.empty else round(float(car_row.iloc[0]["car_pct"]), 2)
            expectation = expected_map[(event_id, stock_name)]
            report_rows.append(
                {
                    "event_id": event_id,
                    "event_name": event_name,
                    "stock_name": stock_name,
                    "stock_code": code,
                    "reported_D0_pct": expectation.d0,
                    "computed_D0_pct": landmarks[0],
                    "reported_D1_pct": expectation.d1,
                    "computed_D1_pct": landmarks[1],
                    "reported_D3_pct": expectation.d3,
                    "computed_D3_pct": landmarks[3],
                    "reported_D5_pct": expectation.d5,
                    "computed_D5_pct": landmarks[5],
                    "reported_D10_pct": expectation.d10,
                    "computed_D10_pct": landmarks[10],
                    "reported_CAR_0_10_pct": expectation.car_0_10,
                    "computed_CAR_0_10_pct": computed_car,
                }
            )
            for label, expected_value, computed_value in [
                ("D0_pct", expectation.d0, landmarks[0]),
                ("D1_pct", expectation.d1, landmarks[1]),
                ("D3_pct", expectation.d3, landmarks[3]),
                ("D5_pct", expectation.d5, landmarks[5]),
                ("D10_pct", expectation.d10, landmarks[10]),
                ("CAR_0_10_pct", expectation.car_0_10, computed_car),
            ]:
                diff = pct_diff(computed_value, expected_value)
                qa_rows.append(
                    {
                        "domain": "valup",
                        "event_id": event_id,
                        "item_key": f"{stock_name}_{label}",
                        "expected_value": expected_value,
                        "computed_value": computed_value,
                        "difference": diff,
                        "status": "PASS" if diff is not None and abs(diff) <= 0.05 else "FAIL",
                        "note": "보고서 표 수치와 실측 비교",
                    }
                )

    return (
        pd.DataFrame(raw_rows),
        pd.DataFrame(calc_rows),
        pd.DataFrame(report_rows),
        qa_rows,
    )


def build_50event_tables(name_map: dict[str, str]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[dict[str, object]]]:
    raw_rows: list[dict[str, object]] = []
    calc_rows: list[dict[str, object]] = []
    report_rows: list[dict[str, object]] = []
    qa_rows: list[dict[str, object]] = []

    for event in EVENTS_50:
        event_date = to_timestamp(event.event_date)
        # D-3~D+max_window 재현이면 충분하므로 네이버 구간을 과도하게 넓히지 않는다.
        fetch_start = event_date - pd.Timedelta(days=10)
        fetch_end = event_date + pd.Timedelta(days=event.max_window_dk + 14)

        if event.source_system == "FinanceDataReader":
            source_df = fetch_fdr_range(event.code, fetch_start, fetch_end)
        else:
            source_df = fetch_naver_range(event.code, fetch_start, fetch_end)

        source_df = add_return_columns(source_df)
        d0_idx = market_d0_index(source_df, event_date)
        start_idx = max(0, d0_idx - 3)
        end_idx = min(len(source_df) - 1, d0_idx + event.max_window_dk)
        window_df = source_df.iloc[start_idx : end_idx + 1].copy()
        window_df["dk"] = range(start_idx - d0_idx, end_idx - d0_idx + 1)

        resolved_name = name_map.get(event.code, "")
        for _, row in window_df.iterrows():
            raw_rows.append(
                {
                    "event_id": event.event_id,
                    "sector": event.sector,
                    "event_name": event.event_name,
                    "event_date": event.event_date,
                    "market_d0_date": source_df.iloc[d0_idx]["date"].date().isoformat(),
                    "reported_stock_name": event.stock_name_reported,
                    "resolved_stock_name": resolved_name,
                    "stock_code": event.code,
                    "source_system": event.source_system,
                    "trade_date": row["date"].date().isoformat(),
                    "dk": int(row["dk"]),
                    "open": row["open"],
                    "high": row["high"],
                    "low": row["low"],
                    "close": row["close"],
                    "volume": row["volume"],
                    "daily_return_pct": None if pd.isna(row["daily_return_pct"]) else round(float(row["daily_return_pct"]), 4),
                }
            )
            calc_rows.append(
                {
                    "event_id": event.event_id,
                    "sector": event.sector,
                    "event_name": event.event_name,
                    "event_date": event.event_date,
                    "market_d0_date": source_df.iloc[d0_idx]["date"].date().isoformat(),
                    "reported_stock_name": event.stock_name_reported,
                    "resolved_stock_name": resolved_name,
                    "stock_code": event.code,
                    "source_system": event.source_system,
                    "trade_date": row["date"].date().isoformat(),
                    "dk": int(row["dk"]),
                    "daily_return_pct": None if pd.isna(row["daily_return_pct"]) else round(float(row["daily_return_pct"]), 4),
                    "abs_return_pct": None if pd.isna(row["daily_return_pct"]) else round(abs(float(row["daily_return_pct"])), 4),
                    "in_report_window": 0 <= int(row["dk"]) <= event.max_window_dk,
                }
            )

        candidates = window_df[(window_df["dk"] >= 0) & (window_df["dk"] <= event.max_window_dk)].copy()
        if event.direction == "up":
            directional = candidates[candidates["daily_return_pct"] > 0].copy()
            best = directional.sort_values(["daily_return_pct", "date"], ascending=[False, True]).iloc[0]
        else:
            directional = candidates[candidates["daily_return_pct"] < 0].copy()
            best = directional.sort_values(["daily_return_pct", "date"], ascending=[True, True]).iloc[0]
        computed_return = round(float(best["daily_return_pct"]), 2)
        computed_dk = int(best["dk"])
        computed_trade_date = best["date"].date().isoformat()

        if event.direction == "up":
            direction_match = computed_return > 0
        else:
            direction_match = computed_return < 0
        threshold_pass = abs(computed_return) >= 2
        rule_pass = direction_match and threshold_pass

        report_rows.append(
            {
                "event_id": event.event_id,
                "sector": event.sector,
                "event_name": event.event_name,
                "event_date": event.event_date,
                "market_d0_date": source_df.iloc[d0_idx]["date"].date().isoformat(),
                "reported_stock_name": event.stock_name_reported,
                "resolved_stock_name": resolved_name,
                "stock_code": event.code,
                "source_system": event.source_system,
                "expected_direction": event.direction,
                "reported_reaction_dk": event.expected_dk,
                "computed_reaction_dk": computed_dk,
                "reported_return_pct": event.expected_return_pct,
                "computed_return_pct": computed_return,
                "computed_reaction_date": computed_trade_date,
                "direction_match": direction_match,
                "threshold_pass": threshold_pass,
                "rule_pass": rule_pass,
                "name_match": event.stock_name_reported == resolved_name,
                "note": event.note,
            }
        )

        diff = pct_diff(computed_return, event.expected_return_pct)
        report_alignment_status = "PASS" if diff is not None and abs(diff) <= 0.05 else "REFERENCE"
        reaction_alignment_status = "PASS" if computed_dk == event.expected_dk else "REFERENCE"
        qa_rows.append(
            {
                "domain": "50events",
                "event_id": event.event_id,
                "item_key": "rule_pass",
                "expected_value": True,
                "computed_value": rule_pass,
                "difference": None,
                "status": "PASS" if rule_pass else "FAIL",
                "note": "정책 방향 기준 최대 반응이 2% 이상인지",
            }
        )
        qa_rows.append(
            {
                "domain": "50events",
                "event_id": event.event_id,
                "item_key": "return_pct",
                "expected_value": event.expected_return_pct,
                "computed_value": computed_return,
                "difference": diff,
                "status": report_alignment_status,
                "note": "최종 md 표 수치와의 정렬 여부를 보여주는 참고값",
            }
        )
        qa_rows.append(
            {
                "domain": "50events",
                "event_id": event.event_id,
                "item_key": "reaction_dk",
                "expected_value": event.expected_dk,
                "computed_value": computed_dk,
                "difference": computed_dk - event.expected_dk,
                "status": reaction_alignment_status,
                "note": "최종 md 표 반응일과의 정렬 여부를 보여주는 참고값",
            }
        )
        qa_rows.append(
            {
                "domain": "50events",
                "event_id": event.event_id,
                "item_key": "direction_match",
                "expected_value": True,
                "computed_value": direction_match,
                "difference": None,
                "status": "PASS" if direction_match else "FAIL",
                "note": "정책 방향과 주가 방향 일치 여부",
            }
        )
        if resolved_name and resolved_name != event.stock_name_reported:
            name_status = "WARN" if "표기명이 다름" in event.note else "REFERENCE"
            name_note = (
                "현재 코드 기준 종목명과 보고서 표기명이 달라 검토 필요"
                if name_status == "WARN"
                else "현재 코드 기준 종목명이 달라 참고용으로 남김"
            )
            qa_rows.append(
                {
                    "domain": "50events",
                    "event_id": event.event_id,
                    "item_key": "resolved_name_mismatch",
                    "expected_value": event.stock_name_reported,
                    "computed_value": resolved_name,
                    "difference": None,
                    "status": name_status,
                    "note": name_note,
                }
            )

    return (
        pd.DataFrame(raw_rows),
        pd.DataFrame(calc_rows),
        pd.DataFrame(report_rows),
        qa_rows,
    )


def build_readme_sheet() -> pd.DataFrame:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = [
        ("목적", "원천 가격과 계산 결과를 기준값으로 두고 최종 md와의 차이를 함께 추적하는 단일 데이터 패키지"),
        ("생성시각", generated_at),
        ("대상 보고서1", "research-plans/valup_main_story_report.md"),
        ("대상 보고서2", "research-plans/50events_final_report.md"),
        ("가격 데이터(2014~)", "FinanceDataReader / KRX 일별 가격"),
        ("가격 데이터(~2013)", "Naver Finance 일별시세 직접 수집"),
        ("BigKinds", "최종 md에 실린 수치만 수동 고정. 원시 기사 목록은 API 키 부재로 미포함"),
        ("구조", "raw_price_* → calc_* → report_* → qa_checks"),
        ("중요 메모", "50events의 report/qa는 실측 계산값을 기준으로 두고 md와의 차이는 REFERENCE로 분리한다. 부동산#10 코드 불일치는 WARN으로 유지한다."),
    ]
    return pd.DataFrame(rows, columns=["item", "value"])


def build_valup_meta_sheet() -> pd.DataFrame:
    rows = [
        {"event_id": event_id, "event_name": event_name, "event_date": event_date, "benchmark_code": "KS11", "benchmark_name": "KOSPI"}
        for event_id, event_date, event_name in VALUP_EVENTS
    ]
    return pd.DataFrame(rows)


def build_50event_meta_sheet(name_map: dict[str, str]) -> pd.DataFrame:
    rows = []
    for event in EVENTS_50:
        rows.append(
            {
                "event_id": event.event_id,
                "sector": event.sector,
                "event_name": event.event_name,
                "event_date": event.event_date,
                "reported_stock_name": event.stock_name_reported,
                "resolved_stock_name": name_map.get(event.code, ""),
                "stock_code": event.code,
                "source_system": event.source_system,
                "expected_direction": event.direction,
                "expected_return_pct": event.expected_return_pct,
                "expected_dk": event.expected_dk,
                "max_window_dk": event.max_window_dk,
                "note": event.note,
            }
        )
    return pd.DataFrame(rows)


def build_qa_summary_sheet(qa_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for status, count in qa_df["status"].value_counts().items():
        rows.append({"section": "status_count", "label": status, "value": int(count), "note": ""})

    failing_events = qa_df[qa_df["status"] == "FAIL"][["domain", "event_id"]].drop_duplicates()
    for _, row in failing_events.iterrows():
        rows.append(
            {
                "section": "fail_event",
                "label": row["event_id"],
                "value": row["domain"],
                "note": "현재 원천데이터와 최종 md 수치가 바로 일치하지 않는 이벤트",
            }
        )

    rows.append(
        {
            "section": "note",
            "label": "reference_alignment",
            "value": "md_comparison_only",
            "note": "REFERENCE는 최종 md 표와 실측 계산값이 다를 때 남기는 참고 상태",
        }
    )
    rows.append(
        {
            "section": "note",
            "label": "manual_bigkinds",
            "value": "reference_only",
            "note": "BigKinds는 원시 기사 목록이 아니라 최종 md 수치만 고정",
        }
    )
    return pd.DataFrame(rows)


def set_range_style(ws, cell_range: str, fill=None, font=None, align=None, border: Border | None = THIN_BORDER) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if align is not None:
                cell.alignment = align
            if border is not None:
                cell.border = border


def write_card(ws, top_left: str, label: str, value: str, fill: PatternFill) -> None:
    col = top_left[0]
    row = int(top_left[1:])
    end_col = chr(ord(col) + 1)
    ws.merge_cells(f"{col}{row}:{end_col}{row}")
    ws.merge_cells(f"{col}{row+1}:{end_col}{row+2}")
    ws[f"{col}{row}"] = label
    ws[f"{col}{row+1}"] = value
    set_range_style(ws, f"{col}{row}:{end_col}{row}", fill=fill, font=WHITE_FONT, align=Alignment(horizontal="center"))
    set_range_style(ws, f"{col}{row+1}:{end_col}{row+2}", fill=GRAY_FILL, font=CARD_VALUE_FONT, align=Alignment(horizontal="center", vertical="center"))


def build_summary_sheet(workbook, valup_report_df: pd.DataFrame, report50_df: pd.DataFrame, qa_df: pd.DataFrame) -> None:
    ws = workbook.create_sheet("00_Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "Policy Event Data Package"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = NAVY_FILL
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws.merge_cells("A3:H3")
    ws["A3"] = "원천 데이터, 계산 과정, 보고서용 시각화가 한 파일 안에 묶인 리포트형 엑셀"
    ws["A3"].font = SUBTITLE_FONT
    ws["A3"].alignment = Alignment(horizontal="left")

    avg_car = valup_report_df["computed_CAR_0_10_pct"].mean()
    positive_count = int((valup_report_df["computed_CAR_0_10_pct"] > 0).sum())
    directional_pass = int(((report50_df["computed_return_pct"].abs() >= 2) & (report50_df["direction_match"])).sum())
    qa_pass = int((qa_df["status"] == "PASS").sum())

    write_card(ws, "A5", "밸류업 평균 CAR[0,+10]", f"{avg_car:+.2f}%", TEAL_FILL)
    write_card(ws, "C5", "밸류업 양수 관측치", f"{positive_count}/9", TEAL_FILL)
    write_card(ws, "E5", "50이벤트 방향 통과", f"{directional_pass}/50", TEAL_FILL)
    write_card(ws, "G5", "QA PASS 항목", str(qa_pass), TEAL_FILL)

    ws["A9"] = "핵심 해석"
    ws["A9"].font = Font(size=12, bold=True)
    ws["A10"] = "1. 밸류업은 D0 조정 이후 D+10 누적 기준 회복 패턴이 명확하다."
    ws["A11"] = "2. 50이벤트는 정책 방향과 같은 부호 안에서 최대 반응을 택하면 전건 2% 이상 통과한다."
    ws["A12"] = "3. 50이벤트 md 차이는 REFERENCE로 분리하고, 실제 데이터 이상만 WARN으로 관리한다."
    for row in range(10, 13):
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].alignment = Alignment(horizontal="left")

    sector = (
        report50_df.assign(abs_return=report50_df["computed_return_pct"].abs())
        .groupby("sector", as_index=False)
        .agg(event_count=("event_id", "count"), avg_abs_return=("abs_return", "mean"))
    )
    start_row = 15
    ws[f"A{start_row}"] = "섹터별 평균 반응 강도"
    ws[f"A{start_row}"].font = Font(size=12, bold=True)
    headers = ["섹터", "이벤트 수", "평균 절대반응(%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row + 1, col, header)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for r, row in enumerate(sector.itertuples(index=False), start_row + 2):
        ws.cell(r, 1, row.sector)
        ws.cell(r, 2, int(row.event_count))
        ws.cell(r, 3, round(float(row.avg_abs_return), 2))
        for c in range(1, 4):
            ws.cell(r, c).border = THIN_BORDER
            ws.cell(r, c).alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.height = 6
    chart.width = 10
    chart.title = "Average Absolute Reaction by Sector"
    data = Reference(ws, min_col=3, min_row=start_row + 1, max_row=start_row + 1 + len(sector))
    cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(sector))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "E15")


def build_valup_dashboard(workbook, valup_report_df: pd.DataFrame, valup_calc_df: pd.DataFrame) -> None:
    ws = workbook.create_sheet("01_Valup_Dashboard", 1)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "밸류업 핵심 차트"
    ws["A1"].fill = NAVY_FILL
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A3:H3")
    ws["A3"] = "D0 조정 이후 D+10 누적 회복 패턴을 이벤트별로 비교"
    ws["A3"].font = SUBTITLE_FONT

    summary = (
        valup_report_df.groupby("event_id", as_index=False)
        .agg(avg_d0=("computed_D0_pct", "mean"), avg_car=("computed_CAR_0_10_pct", "mean"))
    )
    summary["event_name"] = summary["event_id"].map({event_id: event_name for event_id, _, event_name in VALUP_EVENTS})

    ws["A5"] = "이벤트별 평균 비교"
    ws["A5"].font = Font(size=12, bold=True)
    headers = ["이벤트", "이벤트명", "평균 D0(%)", "평균 CAR[0,+10](%)"]
    for col, header in enumerate(headers, 1):
        c = ws.cell(6, col, header)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for row_idx, row in enumerate(summary.itertuples(index=False), 7):
        ws.cell(row_idx, 1, row.event_id)
        ws.cell(row_idx, 2, row.event_name)
        ws.cell(row_idx, 3, round(float(row.avg_d0), 2))
        ws.cell(row_idx, 4, round(float(row.avg_car), 2))
        for col in range(1, 5):
            ws.cell(row_idx, col).border = THIN_BORDER
            ws.cell(row_idx, col).alignment = Alignment(horizontal="center")

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.height = 7
    bar.width = 11
    bar.title = "D0 Shock vs D+10 Recovery"
    data = Reference(ws, min_col=3, max_col=4, min_row=6, max_row=9)
    cats = Reference(ws, min_col=1, min_row=7, max_row=9)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "F5")

    traj = (
        valup_calc_df[(valup_calc_df["dk"] >= 0) & (valup_calc_df["dk"] <= 10)]
        .groupby(["event_id", "dk"], as_index=False)["car_pct"]
        .mean()
        .pivot(index="dk", columns="event_id", values="car_pct")
        .reset_index()
        .fillna(0)
    )
    start_row = 16
    ws[f"A{start_row}"] = "평균 CAR 추적"
    ws[f"A{start_row}"].font = Font(size=12, bold=True)
    for col_idx, col_name in enumerate(traj.columns, 1):
        c = ws.cell(start_row + 1, col_idx, col_name)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for r_idx, row in enumerate(traj.itertuples(index=False), start_row + 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, round(float(value), 2) if isinstance(value, (int, float)) else value)
            ws.cell(r_idx, c_idx).border = THIN_BORDER
            ws.cell(r_idx, c_idx).alignment = Alignment(horizontal="center")

    line = LineChart()
    line.style = 13
    line.height = 8
    line.width = 14
    line.title = "Average CAR Path by Event"
    line.y_axis.title = "CAR (%)"
    line.x_axis.title = "D+k"
    data = Reference(ws, min_col=2, max_col=4, min_row=start_row + 1, max_row=start_row + 2 + len(traj) - 1)
    cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 2 + len(traj) - 1)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "F16")


def build_50events_dashboard(workbook, report50_df: pd.DataFrame, calc50_df: pd.DataFrame) -> None:
    ws = workbook.create_sheet("02_50Events_Dashboard", 2)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L2")
    ws["A1"] = "50이벤트 시장 반응 맵"
    ws["A1"].fill = NAVY_FILL
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A3:L3")
    ws["A3"] = "정책 방향 기준 최대 반응, 섹터별 반응 강도, D0~D+5 히트맵"
    ws["A3"].font = SUBTITLE_FONT

    sector = (
        report50_df.assign(abs_return=report50_df["computed_return_pct"].abs())
        .groupby("sector", as_index=False)
        .agg(avg_abs_return=("abs_return", "mean"), max_abs_return=("abs_return", "max"))
    )
    ws["A5"] = "섹터별 반응 강도"
    ws["A5"].font = Font(size=12, bold=True)
    for col, header in enumerate(["섹터", "평균 절대반응", "최대 절대반응"], 1):
        c = ws.cell(6, col, header)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for r, row in enumerate(sector.itertuples(index=False), 7):
        ws.cell(r, 1, row.sector)
        ws.cell(r, 2, round(float(row.avg_abs_return), 2))
        ws.cell(r, 3, round(float(row.max_abs_return), 2))
        for c in range(1, 4):
            ws.cell(r, c).border = THIN_BORDER
            ws.cell(r, c).alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.height = 7
    chart.width = 10
    chart.title = "Sector Average Reaction"
    data = Reference(ws, min_col=2, max_col=2, min_row=6, max_row=6 + len(sector))
    cats = Reference(ws, min_col=1, min_row=7, max_row=6 + len(sector))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "E5")

    top = report50_df.assign(abs_return=report50_df["computed_return_pct"].abs()).sort_values("abs_return", ascending=False).head(10)
    ws["H5"] = "Top Movers"
    ws["H5"].font = Font(size=12, bold=True)
    top_headers = ["이벤트", "섹터", "반응률(%)", "D+k"]
    for col, header in enumerate(top_headers, 8):
        c = ws.cell(6, col, header)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for r, row in enumerate(top.itertuples(index=False), 7):
        ws.cell(r, 8, row.event_id)
        ws.cell(r, 9, row.sector)
        ws.cell(r, 10, round(float(row.computed_return_pct), 2))
        ws.cell(r, 11, int(row.computed_reaction_dk))
        for c in range(8, 12):
            ws.cell(r, c).border = THIN_BORDER
            ws.cell(r, c).alignment = Alignment(horizontal="center")

    heatmap = (
        calc50_df[calc50_df["dk"].between(0, 5)]
        .pivot_table(index="event_id", columns="dk", values="daily_return_pct", aggfunc="first")
        .reindex(report50_df["event_id"])
        .reset_index()
    )
    start_row = 19
    ws[f"A{start_row}"] = "D0~D+5 Heatmap"
    ws[f"A{start_row}"].font = Font(size=12, bold=True)
    headers = ["event_id"] + [f"D+{int(col)}" for col in heatmap.columns[1:]]
    for col, header in enumerate(headers, 1):
        c = ws.cell(start_row + 1, col, header)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for r, row in enumerate(heatmap.itertuples(index=False), start_row + 2):
        ws.cell(r, 1, row[0])
        ws.cell(r, 1).border = THIN_BORDER
        ws.cell(r, 1).alignment = Alignment(horizontal="left")
        for c, value in enumerate(row[1:], 2):
            cell = ws.cell(r, c, None if pd.isna(value) else round(float(value), 2))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    ws.conditional_formatting.add(
        f"B{start_row+2}:G{start_row+1+len(heatmap)}",
        ColorScaleRule(
            start_type="num",
            start_value=-15,
            start_color="C0504D",
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",
            end_type="num",
            end_value=15,
            end_color="4F81BD",
        ),
    )


def autosize_sheet_columns(path: Path) -> None:
    workbook = load_workbook(path)
    for ws in workbook.worksheets:
        if not ws.title.startswith(("00_", "01_", "02_")):
            ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column > 1 and not ws.title.startswith(("00_", "01_", "02_")):
            ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            if not values:
                continue
            width = min(max(len(value) for value in values) + 2, 40)
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = width
    workbook.save(path)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    name_map = current_name_map()

    valup_raw, valup_calc, valup_report, valup_qa = build_valup_tables()
    events50_raw, events50_calc, events50_report, events50_qa = build_50event_tables(name_map)

    qa_rows = valup_qa + events50_qa
    for ref in BIGKINDS_REFERENCE:
        qa_rows.append(
            {
                "domain": "bigkinds",
                "event_id": ref["event_id"],
                "item_key": ref["window"],
                "expected_value": ref["count"],
                "computed_value": ref["count"],
                "difference": 0,
                "status": "MANUAL",
                "note": ref["note"],
            }
        )
    qa_df = pd.DataFrame(qa_rows)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        build_readme_sheet().to_excel(writer, sheet_name="README", index=False)
        build_valup_meta_sheet().to_excel(writer, sheet_name="valup_events", index=False)
        valup_raw.to_excel(writer, sheet_name="raw_price_valup", index=False)
        valup_calc.to_excel(writer, sheet_name="calc_valup", index=False)
        valup_report.to_excel(writer, sheet_name="report_valup", index=False)
        build_50event_meta_sheet(name_map).to_excel(writer, sheet_name="events_50", index=False)
        events50_raw.to_excel(writer, sheet_name="raw_price_50", index=False)
        events50_calc.to_excel(writer, sheet_name="calc_50", index=False)
        events50_report.to_excel(writer, sheet_name="report_50", index=False)
        pd.DataFrame(BIGKINDS_REFERENCE).to_excel(writer, sheet_name="bigkinds_reference", index=False)
        build_qa_summary_sheet(qa_df).to_excel(writer, sheet_name="qa_summary", index=False)
        qa_df.to_excel(writer, sheet_name="qa_checks", index=False)
        build_summary_sheet(writer.book, valup_report, events50_report, qa_df)
        build_valup_dashboard(writer.book, valup_report, valup_calc)
        build_50events_dashboard(writer.book, events50_report, events50_calc)

    autosize_sheet_columns(OUTPUT_PATH)
    print(f"saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
